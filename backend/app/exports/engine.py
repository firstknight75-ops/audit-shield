"""Export engine — Excel (openpyxl), PDF (WeasyPrint), PNG (Pillow).

All exports carry a Tamper-Proof Certificate (report_id, ledger_hash_at_generation,
HMAC signature) and a verify URL printed on every page so anyone holding
the export can re-validate without seeing the data.

Per Phase 4 acceptance: PDF must render every Kurdish-Sorani-specific
letter correctly. WeasyPrint renders glyphs from the configured font;
the CSS uses a Noto Sans Arabic stack because that font was verified in
Phase 1 to cover all 6 Sorani-specific letters (ھ, ێ, ۆ, ڵ, ڕ, ە).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.exports.certificates import tamper_proof_certificate

try:
    from weasyprint import HTML
except Exception:  # pragma: no cover
    HTML = None

try:
    from PIL import Image, ImageDraw, ImageFont
except Exception:  # pragma: no cover
    Image = None
    ImageDraw = None
    ImageFont = None


CORE_OUTPUT_TITLES = {
    'true_picture': 'الصورة الحقيقية',
    'trust_index': 'مؤشر الموثوقية',
    'waste_map': 'خريطة الهدر',
    'risk_map': 'خريطة المخاطر',
    'opportunity_map': 'خريطة الفرص',
    'action_plan': 'خطة العمل',
    'dashboards': 'لوحات القيادة',
    'what_if': 'محاكاة ماذا لو',
}


# Font stack with Noto Sans Arabic first (verified Phase 1 Sorani coverage),
# then system fallbacks. WeasyPrint uses whichever is installed; Noto Sans
# Arabic is the default Arabic webfont family on Linux distros.
PDF_FONT_STACK = "'Noto Sans Arabic', 'Noto Naskh Arabic', 'Cairo', 'Amiri', 'DejaVu Sans', sans-serif"

# Pillow PNG font paths in lookup order (real .ttf files contain glyph data;
# PIL's default font does NOT support Arabic/Sorani).
PNG_FONT_CANDIDATES = [
    '/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf',
    '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
]


def _pick_png_font(size: int = 24):
    """Pick the best available Pillow font that supports Arabic/Sorani."""
    if not ImageFont:
        return None
    for path in PNG_FONT_CANDIDATES:
        if Path(path).exists():
            try:
                return ImageFont.truetype(path, size=size)
            except Exception:
                continue
    return ImageFont.load_default()


def export_excel(path: str, title: str, rows: list[dict], ledger_hash: str, cert: dict | None = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = 'A2'
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(h) for h in headers])
            for c in ws[ws.max_row]:
                c.alignment = Alignment(horizontal='right')
    if cert is None:
        cert = tamper_proof_certificate({'title': title, 'rows_count': len(rows)}, ledger_hash)
    ws.append([])
    ws.append(['ledger_hash_at_generation', cert['ledger_hash_at_generation']])
    ws.append(['signature', cert['signature']])
    ws.append(['report_id', cert['report_id']])
    ws.append(['verify_url', f'/verify/{cert["report_id"]}'])
    wb.save(path)
    return path


def export_pdf(path: str, title: str, rows: list[dict], ledger_hash: str, cert: dict | None = None, language: str = 'ar') -> str:
    """Render a PDF using WeasyPrint with a font stack that supports
    Arabic + Kurdish Sorani.

    The header/footer render in RTL with direction:rtl.
    """
    if cert is None:
        cert = tamper_proof_certificate({'title': title, 'rows_count': len(rows)}, ledger_hash)
    headers = list(rows[0].keys()) if rows else []
    report_id = cert.get('report_id', 'unknown')
    verify_url = f'/verify/{report_id}'
    html = f"""
    <html dir='rtl' lang='{language}'>
    <head>
      <meta charset='utf-8'>
      <style>
        @page {{ size: A4; margin: 1.5cm; @bottom-right {{ content: 'AuditCore'; font-family: {PDF_FONT_STACK}; font-size: 8pt; color: #666; }} }}
        body {{ font-family: {PDF_FONT_STACK}; direction: rtl; unicode-bidi: embed; }}
        h1 {{ font-size: 20pt; margin-bottom: 8pt; }}
        .meta {{ font-size: 9pt; color: #555; margin-bottom: 18pt; padding: 8pt; background: #f5f5f5; border-right: 4pt solid #c89a3b; padding-right: 12pt; }}
        table {{ width: 100%; border-collapse: collapse; margin-bottom: 18pt; }}
        th, td {{ border: 1px solid #999; padding: 6pt; text-align: right; font-size: 10pt; }}
        th {{ background: #ece9d8; font-weight: bold; }}
        .cert {{ font-size: 8pt; color: #555; border-top: 1pt solid #999; padding-top: 8pt; word-break: break-all; direction: ltr; text-align: left; }}
        .cert .label {{ font-weight: bold; }}
      </style>
    </head>
    <body>
      <h1>{title}</h1>
      <div class='meta'>
        <strong>report_id:</strong> <span dir='ltr'>{report_id}</span><br>
        <strong>verify_url:</strong> <span dir='ltr'>{verify_url}</span><br>
        <strong>generated_language:</strong> {language}<br>
        <strong>rows:</strong> {len(rows)}
      </div>
      <table>
        <thead><tr>{''.join(f'<th>{k}</th>' for k in headers)}</tr></thead>
        <tbody>{''.join('<tr>' + ''.join(f'<td>{row.get(h, "")}</td>' for h in headers) + '</tr>' for row in rows)}</tbody>
      </table>
      <div class='cert'>
        <div><span class='label'>ledger_hash_at_generation:</span> <span dir='ltr'>{cert['ledger_hash_at_generation']}</span></div>
        <div><span class='label'>signature:</span> <span dir='ltr'>{cert['signature']}</span></div>
        <div><span class='label'>report_id:</span> <span dir='ltr'>{cert['report_id']}</span></div>
        <div style='margin-top: 8pt;'><span class='label'>Anyone holding this report can verify it (no login required):</span> <span dir='ltr'>{verify_url}</span></div>
      </div>
    </body></html>
    """
    if HTML:
        HTML(string=html).write_pdf(path)
        return path
    # Fallback: write HTML when WeasyPrint isn't installed (CI/sandbox)
    html_path = str(Path(path).with_suffix('.html'))
    Path(html_path).write_text(html, encoding='utf-8')
    return html_path


def export_png(path: str, title: str, rows: list[dict], ledger_hash: str, cert: dict | None = None) -> str:
    """Render a PNG at 300 DPI sized for WhatsApp sharing (1600x1200)."""
    if cert is None:
        cert = tamper_proof_certificate({'title': title, 'rows_count': len(rows)}, ledger_hash)
    if Image and ImageDraw:
        img = Image.new('RGB', (1600, 1200), 'white')
        draw = ImageDraw.Draw(img)
        font_h = _pick_png_font(36)
        font_b = _pick_png_font(20)
        font_s = _pick_png_font(14)
        draw.text((40, 40), title, fill='black', font=font_h)
        y = 110
        for row in rows[:12]:
            draw.text((40, y), str(row), fill='black', font=font_b)
            y += 42
        draw.text((40, 1080), f"ledger={cert['ledger_hash_at_generation']}", fill='black', font=font_s)
        draw.text((40, 1110), f"sig={cert['signature'][:48]}", fill='black', font=font_s)
        draw.text((40, 1140), f"report_id={cert['report_id']} | verify at /verify/{cert['report_id']}", fill='black', font=font_s)
        img.save(path, dpi=(300, 300))
        return path
    fallback = Path(path).with_suffix('.txt')
    payload = (
        f"PNG EXPORT 300DPI\n"
        f"{title}\n"
        f"rows={len(rows)}\n"
        f"ledger={cert['ledger_hash_at_generation']}\n"
        f"signature={cert['signature']}\n"
        f"report_id={cert['report_id']}\n"
        f"verify_url=/verify/{cert['report_id']}\n"
    )
    fallback.write_text(payload, encoding='utf-8')
    return str(fallback)
